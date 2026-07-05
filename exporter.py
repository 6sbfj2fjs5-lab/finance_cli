from __future__ import annotations

import os
import tempfile
from io import BytesIO
from typing import Iterable

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill

import db

# Ensure Chinese font support for saved chart images
from matplotlib import rcParams, font_manager
import os

# Prefer to register a Chinese font on Windows to avoid missing glyphs in saved charts
_candidates = [
    r"C:\Windows\Fonts\msyh.ttc",
    r"C:\Windows\Fonts\msyh.ttf",
    r"C:\Windows\Fonts\SimHei.ttf",
    r"C:\Windows\Fonts\SimSun.ttc",
    r"C:\Windows\Fonts\MSYH.TTC",
]
for _p in _candidates:
    if os.path.exists(_p):
        try:
            font_manager.fontManager.addfont(_p)
            name = font_manager.FontProperties(fname=_p).get_name()
            rcParams["font.sans-serif"] = [name]
            break
        except Exception:
            continue
else:
    rcParams["font.sans-serif"] = ["DejaVu Sans"]

rcParams["axes.unicode_minus"] = False


def _filter_expenses(months: Iterable[int] | None = None, year: int | None = None) -> list[dict]:
    all_expenses = db.get_expenses()
    if year is None:
        year = 2026

    if months is None:
        months = list(range(1, 13))

    months_set = set(months)
    filtered = []
    for item in all_expenses:
        date_str = item["date"]
        if len(date_str) >= 7:
            expense_year = int(date_str[:4])
            expense_month = int(date_str[5:7])
            if expense_year == year and expense_month in months_set:
                filtered.append(item)
    return filtered


def export_expenses_to_excel(buffer: BytesIO, months: Iterable[int] | None = None, year: int | None = None) -> None:
    expenses = _filter_expenses(months=months, year=year)
    wb = Workbook()
    ws_detail = wb.active
    ws_detail.title = "账单明细"
    ws_detail.append(["ID", "金额", "类型", "支付方式", "分类", "日期", "备注"])
    for item in expenses:
        ws_detail.append([
            item["id"],
            item["amount"],
            item["transaction_type"],
            item["payment_method"],
            item["category"],
            item["date"],
            item["note"] or "",
        ])

    category_stats: dict[str, float] = {}
    type_stats: dict[str, float] = {}
    payment_stats: dict[str, float] = {}
    for item in expenses:
        amt = abs(float(item.get("amount", 0.0)))
        category_stats[item["category"]] = category_stats.get(item["category"], 0.0) + amt
        type_stats[item["transaction_type"]] = type_stats.get(item["transaction_type"], 0.0) + amt
        payment_stats[item["payment_method"]] = payment_stats.get(item["payment_method"], 0.0) + amt

    ws_stats = wb.create_sheet("分类统计")
    ws_stats.append(["分类", "总额"])
    for category, total in sorted(category_stats.items()):
        ws_stats.append([category, total])

    ws_type_stats = wb.create_sheet("类型统计")
    ws_type_stats.append(["类型", "总额"])
    for transaction_type, total in sorted(type_stats.items()):
        ws_type_stats.append([transaction_type, total])

    ws_payment_stats = wb.create_sheet("支付方式统计")
    ws_payment_stats.append(["支付方式", "总额"])
    for payment_method, total in sorted(payment_stats.items()):
        ws_payment_stats.append([payment_method, total])

    ws_chart = wb.create_sheet("统计图")
    ws_chart.append(["图表", "说明"])
    ws_chart.append(["柱状图", "按分类展示金额"])
    ws_chart.append(["收支类型饼图", "按收入/支出分布"])
    ws_chart.append(["支付方式饼图", "按支付方式分布"])

    bar_path = _create_chart_image(category_stats, chart_type="bar", title="分类金额柱状图")
    type_path = _create_chart_image(type_stats, chart_type="pie", title="收支类型分布")
    payment_path = _create_chart_image(payment_stats, chart_type="pie", title="支付方式分布")
    try:
        img_bar = Image(bar_path)
        img_bar.width = 6
        img_bar.height = 3.5
        ws_chart.add_image(img_bar, "A5")

        img_type = Image(type_path)
        img_type.width = 6
        img_type.height = 3.5
        ws_chart.add_image(img_type, "G5")

        img_payment = Image(payment_path)
        img_payment.width = 6
        img_payment.height = 3.5
        ws_chart.add_image(img_payment, "A20")

        for row in ws_detail.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
                cell.font = Font(bold=True)

        wb.save(buffer)
    finally:
        for path in (bar_path, type_path, payment_path):
            if os.path.exists(path):
                os.remove(path)


def _create_chart_image(stats: dict[str, float], chart_type: str, title: str) -> str:
    labels = list(stats.keys())
    values = list(stats.values())
    fig, ax = plt.subplots(figsize=(4, 3))
    if chart_type == "bar":
        ax.bar(labels, values, color="#5B8FF9")
        ax.set_ylabel("金额")
    else:
        if values:
            ax.pie(values, labels=labels, autopct="%1.1f%%", startangle=90)
        else:
            ax.text(0.5, 0.5, "无数据", ha="center", va="center")
    ax.set_title(title)
    ax.set_facecolor("white")
    fig.tight_layout()
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
        path = temp_file.name
    fig.savefig(path, format="png", dpi=150)
    plt.close(fig)
    return path
