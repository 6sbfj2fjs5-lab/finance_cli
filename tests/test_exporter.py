import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

import db
import exporter


class ExporterTests(unittest.TestCase):
    def setUp(self) -> None:
        self.db_path = Path(__file__).resolve().parents[1] / "finance_test.db"
        if self.db_path.exists():
            self.db_path.unlink()

        db.DB_PATH = self.db_path
        db.init_db()
        db.add_expense(15.5, "餐饮", "2026-07-03", "午餐")
        db.add_expense(8.0, "交通", "2026-08-02", "地铁")

    def tearDown(self) -> None:
        if self.db_path.exists():
            self.db_path.unlink()

    def test_export_creates_excel_with_expected_sheets(self) -> None:
        buffer = BytesIO()
        exporter.export_expenses_to_excel(buffer, months=[7], year=2026)
        self.assertGreater(len(buffer.getvalue()), 0)

        buffer.seek(0)
        workbook = load_workbook(buffer)
        self.assertIn("账单明细", workbook.sheetnames)
        self.assertIn("分类统计", workbook.sheetnames)
        self.assertIn("统计图", workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
